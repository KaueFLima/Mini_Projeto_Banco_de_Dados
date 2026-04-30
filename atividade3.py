from sqlalchemy import text
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from decorator_tempo import medir_tempo
from LGPD import engine

HEADER_FILL = PatternFill('solid', start_color='4472C4')
HEADER_FONT = Font(bold=True, color='FFFFFF', name='Arial')
CELL_FONT   = Font(name='Arial')
CENTER      = Alignment(horizontal='center')

def _cabecalho(ws, headers):
    ws.append(headers)
    for cell in ws[1]:
        cell.fill      = HEADER_FILL
        cell.font      = HEADER_FONT
        cell.alignment = CENTER

def _ajustar_colunas(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value), default=10)
        ws.column_dimensions[col[0].column_letter].width = max_len + 4

@medir_tempo
def atividade3():
    with engine.connect() as conn:
        rows = conn.execute(text("SELECT nome, cpf FROM usuarios;")).fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "todos"
    _cabecalho(ws, ['nome', 'cpf'])
    for row in rows:
        ws.append([row.nome, row.cpf])
        for cell in ws[ws.max_row]:
            cell.font = CELL_FONT
    _ajustar_colunas(ws)
    wb.save("todos.xlsx")
    print(f"todos.xlsx — {len(rows)} registro(s)")

if __name__ == '__main__':
    print("Atividade 3 — todos.xlsx (sem anonimização):\n")
    atividade3()
