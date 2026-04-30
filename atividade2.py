from sqlalchemy import text
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from decorator_tempo import medir_tempo
from LGPD import engine
from atividade1 import anonimizar_nome, anonimizar_cpf, anonimizar_email, anonimizar_telefone

HEADER_FILL = PatternFill('solid', start_color='4472C4')
HEADER_FONT = Font(bold=True, color='FFFFFF', name='Arial')
CELL_FONT   = Font(name='Arial')
CENTER      = Alignment(horizontal='center')

def _cabecalho(ws, headers):
    ws.append(headers)
    for cell in ws[1]:
        cell.fill      = HEADER_FILL
        cell.font      = HEADER_FONT
        cell.alignment = CENTER

def _ajustar_colunas(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value), default=10)
        ws.column_dimensions[col[0].column_letter].width = max_len + 4

@medir_tempo
def atividade2():
    with engine.connect() as conn:
        rows = conn.execute(
            text("SELECT * FROM usuarios WHERE EXTRACT(YEAR FROM data_nascimento) = 1998;")
        ).fetchall()

    por_ano = defaultdict(list)
    for row in rows:
        d = dict(row._mapping)
        d['nome']     = anonimizar_nome(d['nome'])
        d['cpf']      = anonimizar_cpf(d['cpf'])
        d['email']    = anonimizar_email(d['email'])
        d['telefone'] = anonimizar_telefone(d['telefone'])
        por_ano[d['data_nascimento'].year].append(d)

    headers = ['id', 'nome', 'cpf', 'email', 'telefone', 'data_nascimento']
    for ano, registros in sorted(por_ano.items()):
        wb = Workbook()
        ws = wb.active
        ws.title = str(ano)
        _cabecalho(ws, headers)
        for r in registros:
            ws.append([r['id'], r['nome'], r['cpf'], r['email'],
                       r['telefone'], str(r['data_nascimento'])])
            for cell in ws[ws.max_row]:
                cell.font = CELL_FONT
        _ajustar_colunas(ws)
        wb.save(f"{ano}.xlsx")
        print(f"{ano}.xlsx — {len(registros)} registro(s)")

if __name__ == '__main__':
    print("Atividade 2 — Arquivos por ano de nascimento:\n")
    atividade2()